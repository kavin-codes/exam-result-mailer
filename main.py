"""
🚨 Disclaimer:

This automation script was originally developed for retrieving exam results from Anna University. 
For security and privacy reasons, the actual university website URL has been removed. 
You can replace `TARGET_WEBSITE_URL` with the intended URL when running in your own secure environment.
TARGET_WEBSITE_URL= anna university resukt webiste .
Note: Always ensure you have proper authorization before automating logins or scraping any website.
"""



from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import datetime
from PIL import Image, ImageOps, ImageFilter, ImageEnhance
import easyocr
import openpyxl

# ✅ Load credentials from Excel (just for login)
wb = openpyxl.load_workbook("sele.xlsx")
sheet = wb.active

credentials = []
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    username, password = row[0], row[1]
    if username and password:
        if isinstance(password, datetime.datetime):
            password = password.strftime("%d-%m-%Y")
        else:
            password = str(password)
        credentials.append((str(username), password))

if not credentials:
    print("⚠️ No credentials found in Excel. Stopping process.")
    exit()

print(f"✅ Loaded {len(credentials)} credentials.")

# ✅ Setup Chrome (update path to your local chromedriver)
chrome_driver_path = "PATH_TO_YOUR_CHROMEDRIVER"  # <-- Replace this with your chromedriver path insall based on your chrome version
service = Service(chrome_driver_path)
driver = webdriver.Chrome(service=service)
driver.maximize_window()

# ✅ Open the target website (replace with your URL)
target_url = "TARGET_WEBSITE_URL"  # <-- Replace this with your website URL||  
driver.get(target_url)

wait = WebDriverWait(driver, 10)
reader = easyocr.Reader(['en'], gpu=False)
max_retries = 5

# ---------- LOOP ----------
for idx, (username, password) in enumerate(credentials, start=1):
    print(f"\n👤 Processing User {idx}: {username} / {password}")
    login_success = False

    for attempt in range(max_retries):
        try:
            print(f"   🔄 Login attempt {attempt+1}...")

            register_input = wait.until(EC.presence_of_element_located((By.ID, "register_no")))
            register_input.clear()
            register_input.send_keys(username)

            dob_input = wait.until(EC.presence_of_element_located((By.ID, "dob")))
            dob_input.clear()
            dob_input.send_keys(password)

            # ✅ CAPTCHA Screenshot
            captcha_element = wait.until(EC.presence_of_element_located((By.XPATH, "(//img[@id='capt'])[2]")))
            captcha_image_path = f"captcha_{username}.png"
            captcha_element.screenshot(captcha_image_path)

            # ✅ Preprocess CAPTCHA
            img = Image.open(captcha_image_path)
            img = img.convert("L")
            img = img.filter(ImageFilter.MedianFilter(size=3))
            img = ImageOps.autocontrast(img, cutoff=2)
            img = ImageEnhance.Sharpness(img).enhance(1.5)
            img.save(captcha_image_path)

            # ✅ OCR
            results = reader.readtext(captcha_image_path, detail=0,
                                      allowlist="abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789")
            captcha_text = ''.join(results).replace(' ', '')
            captcha_text = ''.join(filter(str.isalnum, captcha_text))
            print("   🧠 CAPTCHA Text:", captcha_text)

            captcha_input = driver.find_element(By.ID, "security_code_student")
            captcha_input.clear()
            captcha_input.send_keys(captcha_text)

            # ✅ Login
            login_button = driver.find_element(By.XPATH, "(//input[@id='gos'])[2]")
            login_button.click()
            time.sleep(5)

            if len(driver.find_elements(By.ID, "formExamResults")) > 0:
                print("   ✅ Login successful!")
                login_success = True
                break
            else:
                print("   ❌ Wrong CAPTCHA, retrying...")
                driver.execute_script("""
                var img = document.querySelector('#capt');
                img.src = img.src.split('?')[0] + '?rand=' + Math.random();
                """)
                time.sleep(2)

        except Exception as e:
            print("   ⚠️ Error during login:", e)
            driver.refresh()
            time.sleep(3)

    # ---------- AFTER LOGIN ----------
    if login_success:
        try:
            driver.execute_script("document.getElementById('formExamResults').submit();")
            time.sleep(4)

            rows = driver.find_elements(By.XPATH, "//table[@id='resulttable']//tr")
            screenshot_path = f"exam_result_{username}.png"
            driver.save_screenshot(screenshot_path)
            print(f"   📸 Screenshot saved at: {screenshot_path}")

            print("\n   📘 Exam Results:")
            for row in rows:
                cols = row.find_elements(By.TAG_NAME, "td")
                data = [col.text.strip() for col in cols]
                if data and len(data) >= 5:
                    print("   ", data)

        except Exception as e:
            print("   ⚠️ Could not extract results:", e)

    else:
        print(f"   ❌ Failed to login for user {username} after {max_retries} attempts.")

print("\n🚀 Process finished for all users. Browser stays open at the last user’s results page.")
